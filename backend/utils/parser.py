import os
import csv
from pypdf import PdfReader
from docx import Document
import openpyxl

def parse_txt(file_path: str) -> str:
    """Extract text from a TXT file."""
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()

def parse_pdf(file_path: str) -> str:
    """Extract text from a PDF file."""
    try:
        reader = PdfReader(file_path)
        text = []
        for page in reader.pages:
            t = page.extract_text()
            if t:
                text.append(t)
        return "\n".join(text)
    except Exception as e:
        return f"Error parsing PDF: {str(e)}"

def parse_docx(file_path: str) -> str:
    """Extract text from a DOCX file."""
    try:
        doc = Document(file_path)
        text = [p.text for p in doc.paragraphs]
        return "\n".join(text)
    except Exception as e:
        return f"Error parsing DOCX: {str(e)}"

def parse_csv(file_path: str) -> str:
    """Extract text summary from a CSV file."""
    try:
        rows = []
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            reader = csv.reader(f)
            for idx, row in enumerate(reader):
                if idx < 100: # Limit to first 100 rows
                    rows.append(", ".join(row))
                else:
                    rows.append("... [truncated]")
                    break
        return "\n".join(rows)
    except Exception as e:
        return f"Error parsing CSV: {str(e)}"

def parse_excel(file_path: str) -> str:
    """Extract text summary from an Excel spreadsheet."""
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        output = []
        for sheet_name in wb.sheetnames[:3]: # Limit to first 3 sheets
            sheet = wb[sheet_name]
            output.append(f"Sheet: {sheet_name}")
            for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if r_idx < 50: # Limit to 50 rows per sheet
                    row_str = ", ".join([str(val) if val is not None else "" for val in row])
                    output.append(row_str)
                else:
                    output.append("... [truncated]")
                    break
        return "\n".join(output)
    except Exception as e:
        return f"Error parsing Excel: {str(e)}"

def parse_file(file_path: str) -> str:
    """Extract text from supported document types based on extension."""
    if not os.path.exists(file_path):
        return f"Error: File '{file_path}' does not exist."
        
    ext = os.path.splitext(file_path)[1].lower()
    
    if ext == ".txt":
        return parse_txt(file_path)
    elif ext == ".pdf":
        return parse_pdf(file_path)
    elif ext in [".docx", ".doc"]:
        return parse_docx(file_path)
    elif ext == ".csv":
        return parse_csv(file_path)
    elif ext in [".xlsx", ".xls"]:
        return parse_excel(file_path)
    else:
        return f"Unsupported file format: {ext}"
